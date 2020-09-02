#!/usr/bin/env python
# -*- coding: utf-8 -*-
___author___ = 'Steven Emerson'
___date___ = '2020/8/31 10:14'

import datetime

import yaml
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted

ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

